from openpyxl import load_workbook
from mymath import *
import numpy as np

snames = ['stress', 'mpa', 'напряжение', 'усилие']
enames = ['strain', 'extension', 'deformation', '%', 'деформация']


def mk_df(file):
    global snames, enames, dtypes
    stress = {}
    extension = {}
    wb = load_workbook(file)

    for sheetname in wb.sheetnames:
        offset = 1
        while type(wb[sheetname]['A' + str(offset)].value) not in dtypes:

            offset += 1
        print((wb[sheetname]['A' + str(offset)].value) not in dtypes)
        print(type(wb[sheetname]['A' + str(offset)].value) not in dtypes)
        for row in (wb[sheetname]['A1:Z' + str(offset)]):
            for cell in row:
                if type(cell.value) is str:
                    for name in snames + enames:
                        if name in cell.value.lower():
                            if name in snames:
                                stress_col = cell.column_letter
                            elif name in enames:
                                ext_col = cell.column_letter
        print(stress_col, ext_col)
        stress[sheetname] = []
        extension[sheetname] = []
        while type(wb[sheetname][stress_col + str(offset)].value) in dtypes:
            stress[sheetname].append(wb[sheetname][stress_col + str(offset)].value)
            extension[sheetname].append(wb[sheetname][ext_col + str(offset)].value)
            offset += 1

    return stress, extension



def _reduce2(l):
    ans = []
    for i in range(0, len(l), 2)[::-1]:
        ans.append(l[i])
    return ans

def less_lenght(l, lenght = 1000, reduce=_reduce2):
    if len(l) > lenght:
        return less_lenght(reduce(l), lenght)
    else:
        return l